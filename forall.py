import numpy as np
from skimage import io, draw
import matplotlib.pyplot as plt
import cv2 as cv
import numpy as np
import math
import os
import openpyxl

folder_path = 'C:/Users/sravi/PycharmProject/formal_proj/wetransfer_droplet-image-analysis_2023-01-21_0550/Images_analysis/'

file_list = [f for f in os.listdir(folder_path) if f.endswith('.tif')]
image_list = []
red_list = []
green_list = []

# generate correct list of images
for i in range(int(len(file_list)/5)):
    for j in range(5):
        if(j==1):
            green_list.append(file_list[i*5+j])
        if(j==2):
            red_list.append(file_list[i*5+j])
        if(j==3):
            image_list.append(file_list[i*5+j])

print(len(file_list))
print(len(image_list))
droplets_all=[]  #new vector pushed for every image after matching
total_drop=[]  #coord change after every image
distance_all=[]
xmatch=[]
ymatch=[]
droplets2new=[]
droplets2=[]

for i, file_name in enumerate(image_list):
    if(i==0):
        img = cv.imread(os.path.join(folder_path, file_name))
        
        gray= cv.cvtColor(img, cv.COLOR_BGR2GRAY)
        ret, bw_img = cv.threshold(gray, 15, 255, cv.THRESH_BINARY)
        img_canny = cv.Canny(bw_img, 50, 150)  #to enhance edges

        # Get the dimensions of the original image
        height, width, channels = img.shape        
        # Calculate the dimensions of the large rectangular region
        factor = 2
        rect_height = int(height / factor)
        rect_width = int(width / factor)        
        # Calculate the top-left and bottom-right coordinates of the large rectangular region
        top_left = (int(width / 2 - rect_width / 2), int(height / 2 - rect_height / 2))
        bottom_right = (int(width / 2 + rect_width / 2), int(height / 2 + rect_height / 2))        
        # Create a binary mask image of the same dimensions as the original image
        mask = np.zeros((height, width), dtype=np.uint8)        
        # Set the pixels inside the large rectangular region to 255 (white) and the rest of the pixels to 0 (black)
        cv.rectangle(mask, top_left, bottom_right, 255, -1)        
        # Find contours in the large rectangular region
        masked_image = cv.bitwise_and(img_canny, img_canny, mask=mask)

        contours, hierarchy = cv.findContours(masked_image, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)  #detect edges

        droplets=[]

        # Fit a circle to the detected edges using the minEnclosingCircle() function
        for contour in contours:
            (x,y), radius = cv.minEnclosingCircle(contour)
            center = (int(x), int(y))
            radius = int(radius)
            #center_subpixel = (round(x,2), round(y,2))
            center_subpixel_x = round(x,2)
            center_subpixel_y = round(y,2)
            
            # Mark the detected circle and center point on the output image
            area = np.pi * radius**2
            if area > 900:
                droplets.append((center_subpixel_x, center_subpixel_y, radius))
        total_drop=droplets
        a=len(droplets)
        droplets_all.append(droplets)
    else:
        img = cv.imread(os.path.join(folder_path, file_name))
        # cv.imshow("img",img)
        gray= cv.cvtColor(img, cv.COLOR_BGR2GRAY)
        ret, bw_img = cv.threshold(gray, 15, 255, cv.THRESH_BINARY)
        img_canny = cv.Canny(bw_img, 50, 150)  #to enhance edges

        # Get the dimensions of the original image
        height, width, channels = img.shape        
        # Calculate the dimensions of the large rectangular region
        factor = 2
        rect_height = int(height / factor)
        rect_width = int(width / factor)        
        # Calculate the top-left and bottom-right coordinates of the large rectangular region
        top_left = (int(width / 2 - rect_width / 2), int(height / 2 - rect_height / 2))
        bottom_right = (int(width / 2 + rect_width / 2), int(height / 2 + rect_height / 2))        
        # Create a binary mask image of the same dimensions as the original image
        mask = np.zeros((height, width), dtype=np.uint8)        
        # Set the pixels inside the large rectangular region to 255 (white) and the rest of the pixels to 0 (black)
        cv.rectangle(mask, top_left, bottom_right, 255, -1)        
        # Find contours in the large rectangular region
        masked_image = cv.bitwise_and(img_canny, img_canny, mask=mask)

        contours, hierarchy = cv.findContours(masked_image, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)  #detect edges

        droplets2.clear()
        # j=0

        # Fit a circle to the detected edges using the minEnclosingCircle() function
        for contour in contours:
            (x,y), radius = cv.minEnclosingCircle(contour)
            center = (int(x), int(y))
            radius = int(radius)
            center_subpixel_x = round(x,2)
            center_subpixel_y = round(y,2)
            area = np.pi * radius**2
            if area > 900:
                droplets2.append((center_subpixel_x, center_subpixel_y, radius))
        droplets2new=[]
        dist=[]
        #now we have droplets and droplets2. want to match everything and then push new vector into droplets_all
        for d, drop in enumerate(droplets_all[i-1]):
            minD=3.0
            for drop2 in droplets2:
                distance=math.sqrt(pow((drop2[0]-drop[0]), 2)+pow((drop2[1]-drop[1]), 2)) 
                xmatch=-1
                ymatch=-1
                if distance<minD:
                    minD=distance
                    xmatch=drop2[0]
                    ymatch=drop2[1]
                    break
            droplets2new.append((xmatch,ymatch, drop2[2]))
            if(distance<4):
                dist.append(distance)  #place in right position
            else:
                dist.append(-1)   #append assumes match found for every droplet so no gaps in id
      
        droplets_all.append(droplets2new)
        distance_all.append(dist)

workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the data in droplets_all to the worksheet
for i, droplet in enumerate(droplets_all, start=1):
    worksheet.cell(row=1, column=i+1, value=i)
    for j, coord in enumerate(droplet, start=2):
        x, y, r = coord
        worksheet.cell(row=j, column=1, value=j-1)
        worksheet.cell(row=j, column=i+1, value="({}, {})".format(x, y))

# Save the workbook
workbook.save('droplets_all.xlsx')

# Create a new Excel workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write the data in distance_all to the worksheet
for i, dist in enumerate(distance_all, start=1):
    worksheet.cell(row=1, column=i+1, value="{0}-{1}".format(i, i+1))
    for j, d in enumerate(dist, start=2):
        worksheet.cell(row=j, column=1, value=j-1)
        worksheet.cell(row=j, column=i+1, value=d)

# Save the workbook
workbook.save('distance_all.xlsx')

#valid_droplets contains only information of droplets that stay in the frame throughout the video 
valid_droplets = [[] for _ in range(len(droplets_all[0]))]

for i, coords in enumerate(zip(*droplets_all)):
    if (-1, -1) not in coords:
        for j, c in enumerate(coords):
            valid_droplets[j].append(c)

print(len(valid_droplets))

wb = openpyxl.Workbook()
sheet = wb.active

for i, droplet in enumerate(valid_droplets):
    for j, coord in enumerate(droplet):
        sheet.cell(row=j+1, column=i+1, value=str(coord))

wb.save('valid_droplets.xlsx')

#valid_distance generation and writing to excel
# valid_distance = []
# for distance in distance_all:
#     if -1 not in distance:
#         valid_distance.append(distance)

# print(len(valid_distance))

import matplotlib.pyplot as plt

# Time interval between images in seconds
interval = 120

# Create a new figure
fig, ax = plt.subplots()

# Plot distance data for each entity with a separate color, excluding any distance value of -1
for i, distance in enumerate(distance_all):
    filtered_distance = [d for d in distance if d != -1]
    time = [t * interval for t in range(len(filtered_distance))]
    ax.plot(time, filtered_distance, label=f'Drop {i+1}')

# Set labels for x and y axes
ax.set_xlabel('Time (sec)')
ax.set_ylabel('Distance (m)')

# Add a legend to the plot
ax.legend()

# Show the plot
plt.show()

#sep graphs using valid_distance

# Create folder for saving graphs
# if not os.path.exists('graphs'):
#     os.makedirs('graphs')

# # Iterate over each distance vector
# for i, drop in enumerate(distance_all):
#     # Calculate time values for each distance vector
#     time = [t*120 for t in range(len(drop))]
#     # Plot distance-time graph for the drop
#     plt.plot(time, drop, label=f"Drop {i+1}")
#     plt.xlabel("Time (sec)")
#     plt.ylabel("Distance (mm)")
#     plt.title(f"Drop {i+1} Distance-Time Graph")
#     plt.legend()
#     # Save the graph image in the "graphs" folder
#     plt.savefig(f"graphs/drop_{i+1}_graph.png")
#     # Clear the plot for the next iteration
#     plt.close(fig)
#     plt.clf()

################fluoresence part##############################

intensities_all = []
for i, filename in enumerate(red_list):
    # Load image
    image_name = os.path.join(folder_path, filename)
    image = io.imread(image_name)

    # Initialize intensities vector for this image
    intensities = np.zeros(len(droplets_all[i]))

    # Loop over each droplet in this image
    for j, (x, y, r) in enumerate(droplets_all[i]):
        # Create circular mask for this droplet
        mask = np.zeros(image.shape, dtype=bool)
        rr, cc = draw.disk((int(round(y)), int(round(x))), r, shape=image.shape)
        mask[rr, cc] = True

        # Compute mean intensity within the mask
        pixels_in_droplet = image[mask]
        mean_intensity = np.mean(pixels_in_droplet)

        # Store intensity in intensities vector
        intensities[j] = mean_intensity

    # Append intensities vector to intensities_all
    intensities_all.append(intensities)

def isInside(circle_x, circle_y, rad, x, y):
    if ((x - circle_x) * (x - circle_x) + (y - circle_y) * (y - circle_y) <= rad * rad):
        return True
    else:
        return False

#only generating multiple-cell containing index in first image
image = cv.imread('Series1_29Nov_t001_c002.tif')

# Convert image to Grayscale
gray = cv.cvtColor(image, cv.COLOR_BGR2GRAY)

# Threshold Binary
ret,thresh1 = cv.threshold(gray, 15, 255, cv.THRESH_BINARY)

# Detect contours
contours, hierarchy = cv.findContours(thresh1, cv.RETR_TREE, cv.CHAIN_APPROX_NONE)

# Draw contours on grayscale image
img2 = gray.copy()
cv.drawContours(img2, contours, -1, (255,255,0), 3)

# Create list to store indices of droplets with multiple cells
multi_cell_drops = []
cell_drops = []      # Initialize vector for droplets with one or more cells
no_cell_drops = []   # Initialize vector for droplets with no cells

# Loop through each droplet in current image
for j, droplet in enumerate(valid_droplets[0]):
    x, y, r = droplet
    
    # Loop through each contour and check if it's inside the droplet
    cell_count = 0
    for contour in contours:
        for point in contour:
            if isInside(x, y, r, point[0][0], point[0][1]):
                cell_count += 1
                break  # Only count the contour as one cell if any point is inside the droplet
    if cell_count>1:
        multi_cell_drops.append(j)
    if cell_count >= 1:
        cell_drops.append(j)    # Append index of droplet with one or more cells to cell_drops[]
    elif cell_count == 0:
        no_cell_drops.append(j) # Append index of droplet with no cells to no_cell_drops[]


print("Droplets with cells in:", len(cell_drops))
print("Droplets with no cells in:", len(no_cell_drops))

import matplotlib.pyplot as plt
import matplotlib.cm as cm

# Define time intervals between images
time_interval = 120  # seconds

# Compute time stamps for each image
timestamps = [i * time_interval for i in range(len(image_list))]


# Set colors for cell_drops and no_cell_drops
color_cell = 'blue'
color_no_cell = 'red'

# Plot intensity-time graph for each droplet
for i in range(len(valid_droplets[0])):
    # Check if this droplet is in cell_drops
    if i in cell_drops:
        color = color_cell
    # Check if this droplet is in no_cell_drops
    elif i in no_cell_drops:
        color = color_no_cell
    else:
        continue  # Skip if droplet is not in cell_drops or no_cell_drops

    # Plot intensities up to 5880 seconds for this droplet
    intensities = [intensities_all[j][i] for j in range(len(image_list)) if j*time_interval <= 5880]
    timestamps = [j*time_interval for j in range(len(intensities))]
    plt.plot(timestamps, intensities, color=color, linewidth=0.5)

# Add legend and labels
plt.legend()
plt.xlabel('Time (s)')
plt.ylabel('Intensity')
plt.show()